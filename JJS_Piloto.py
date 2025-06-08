import sqlite3
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import csv
import pandas as pd
import plotly.express as px
import shutil
import os
from PIL import Image, ImageTk
from jj_login_module import abrir_login
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl import load_workbook
import plotly.io as pio
pio.kaleido.scope.default_format = "pdf"
user = None

# ====== BANCO DE DADOS ======
conn = sqlite3.connect("estoque.db")
cursor = conn.cursor()

cursor.execute("""
CREATE TABLE IF NOT EXISTS produtos (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    codigo TEXT,
    lote TEXT,
    ordem_producao TEXT,
    quantidade_total INTEGER,
    quantidade_defeituosa INTEGER,
    tipo_defeito TEXT,
    classe_defeito TEXT,
    tag TEXT,
    turno TEXT,
    data_reprova TEXT,
    tecnico TEXT,
    comentarios TEXT,
    foto TEXT
)
""")
cursor.execute("""
CREATE TABLE IF NOT EXISTS movimentacoes (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    produto_id INTEGER,
    tipo TEXT,
    quantidade INTEGER,
    data TEXT,
    FOREIGN KEY(produto_id) REFERENCES produtos(id)
)
""")

# ====== TABELA DE LOG DE OPERAÇÕES ======
cursor.execute("""
CREATE TABLE IF NOT EXISTS log_operacoes (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    operacao TEXT,
    codigo TEXT,
    lote TEXT,
    usuario TEXT,
    data_hora TEXT
)
""")
conn.commit()

conn.commit()

# ====== ÍNDICES PARA OTIMIZAÇÃO ======
cursor.execute("CREATE INDEX IF NOT EXISTS idx_codigo ON produtos(codigo)")
cursor.execute("CREATE INDEX IF NOT EXISTS idx_lote ON produtos(lote)")
cursor.execute("CREATE INDEX IF NOT EXISTS idx_data ON produtos(data_reprova)")
cursor.execute("CREATE INDEX IF NOT EXISTS idx_turno ON produtos(turno)")
cursor.execute("CREATE INDEX IF NOT EXISTS idx_tipo_defeito ON produtos(tipo_defeito)")
cursor.execute("CREATE INDEX IF NOT EXISTS idx_classe_defeito ON produtos(classe_defeito)")
conn.commit()


# ====== FUNÇÕES ======
def selecionar_foto():
    caminho = filedialog.askopenfilename(filetypes=[("Imagens", "*.png;*.jpg;*.jpeg;*.bmp")])
    if caminho:
        foto_path.set(caminho)
        messagebox.showinfo("Foto selecionada", f"Foto selecionada com sucesso:\n{caminho}")

def salvar_produto():
    try:
        foto_destino = ""
        if foto_path.get():
            os.makedirs("fotos_reprovas", exist_ok=True)
            nome_arquivo = os.path.basename(foto_path.get())
            foto_destino = os.path.join("fotos_reprovas", nome_arquivo)
            shutil.copy(foto_path.get(), foto_destino)

        dados = (
            codigo_entry.get(), lote_entry.get(), ordem_entry.get(), int(qtd_total_entry.get()),
            int(qtd_defeituosa_entry.get()), tipo_defeito_entry.get(), classe_defeito_entry.get(),
            tag_entry.get(), turno_entry.get(), data_reprova_entry.get(), tecnico_entry.get(),
            comentarios_entry.get(), foto_destino
        )
        cursor.execute("""
            INSERT INTO produtos (
                codigo, lote, ordem_producao, quantidade_total, quantidade_defeituosa,
                tipo_defeito, classe_defeito, tag, turno, data_reprova, tecnico, comentarios, foto
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, dados)
        conn.commit()

        # Registrar log de inserção
        cursor.execute("""
            INSERT INTO log_operacoes (operacao, codigo, lote, usuario, data_hora)
            VALUES (?, ?, ?, ?, ?)
        """, (
            "inserção",
            codigo_entry.get(),
            lote_entry.get(),
            user,  # <-- essa variável precisa existir no escopo atual
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ))
        conn.commit()

        messagebox.showinfo("Sucesso", "Produto cadastrado com sucesso!")
        limpar_campos()
        atualizar_tabela()
        foto_path.set("")

    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao salvar produto: {e}")


def registrar_movimentacao():
    try:
        produto_id = int(mov_produto_id.get())
        tipo = mov_tipo.get()
        quantidade = int(mov_quantidade.get())
        data = datetime.now().strftime("%Y-%m-%d")

        cursor.execute(
            "INSERT INTO movimentacoes (produto_id, tipo, quantidade, data) VALUES (?, ?, ?, ?)",
            (produto_id, tipo, quantidade, data)
        )
        conn.commit()

        # Buscar informações do produto para log
        cursor.execute("SELECT * FROM produtos WHERE id = ?", (produto_id,))
        produto_info = cursor.fetchone()
        if produto_info:
            cursor.execute("""
                INSERT INTO log_operacoes (operacao, codigo, lote, usuario, data_hora)
                VALUES (?, ?, ?, ?, ?)
            """, (
                f"movimentação ({tipo})",
                produto_info[1],  # código
                produto_info[2],  # lote
                user,             # variável recebida do login
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ))
            conn.commit()

        messagebox.showinfo("Sucesso", "Movimentação registrada!")
        mov_produto_id.delete(0, tk.END)
        mov_quantidade.delete(0, tk.END)
        atualizar_tabela_mov()

    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao registrar movimentação: {e}")

def limpar_campos():
    for entry in entradas:
        entry.delete(0, tk.END)

def atualizar_tabela(filtro=""):
    for row in tabela.get_children():
        tabela.delete(row)
    query = "SELECT * FROM produtos"
    if filtro:
        query += f" WHERE codigo LIKE '%{filtro}%' OR lote LIKE '%{filtro}%' OR tecnico LIKE '%{filtro}%'"
    cursor.execute(query)
    for row in cursor.fetchall():
        row = list(row)
        try:
            row[10] = datetime.strptime(row[10], "%Y-%m-%d").strftime("%d/%m/%Y")
        except:
            pass
        tabela.insert("", tk.END, values=row)

def atualizar_tabela_mov():
    for row in tabela_mov.get_children():
        tabela_mov.delete(row)
    cursor.execute("SELECT * FROM movimentacoes")
    for row in cursor.fetchall():
        row = list(row)
        try:
            row[4] = datetime.strptime(row[4], "%Y-%m-%d").strftime("%d/%m/%Y")
        except:
            pass
        tabela_mov.insert("", tk.END, values=row)

def exportar_produtos_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Arquivos Excel", "*.xlsx")],
        title="Salvar Produtos como Excel"
    )
    if not file_path:
        return

    dados = [tabela.item(row)["values"] for row in tabela.get_children()]
    if not dados:
        messagebox.showinfo("Sem dados", "Não há registros para exportar.")
        return

    colunas = ["ID", "Código", "Lote", "Ordem", "Qtd Total", "Qtd Defeituosa",
               "Tipo Defeito", "Classe Defeito", "Tag", "Turno", "Data Reprova",
               "Técnico", "Comentários", "Foto"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="EB1700")
    header_align = Alignment(horizontal="center")

    for col_index, titulo in enumerate(colunas, start=1):
        cell = ws.cell(row=1, column=col_index, value=titulo)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_index, linha in enumerate(dados, start=2):
        for col_index, valor in enumerate(linha, start=1):
            ws.cell(row=row_index, column=col_index, value=valor)

    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        col_letra = get_column_letter(col[0].column)
        ws.column_dimensions[col_letra].width = max_len + 2

    wb.save(file_path)
    messagebox.showinfo("Exportado", "Produtos exportados com sucesso em Excel!")

import plotly.io as pio
from tkinter import filedialog, messagebox

def salvar_grafico_jpeg(fig, nome_arquivo):
    try:
        # Exibe o gráfico
        fig.show()

        # Pergunta se o usuário quer salvar como JPEG
        if messagebox.askyesno("Exportar JPEG", "Deseja exportar este gráfico como JPEG?"):
            file_path = filedialog.asksaveasfilename(
                defaultextension=".jpeg",
                filetypes=[("Imagem JPEG", "*.jpeg")],
                title="Salvar Gráfico como JPEG",
                initialfile=f"{nome_arquivo}.jpeg"
            )
            if file_path:
                pio.write_image(fig, file_path, format="jpeg", engine="kaleido")
                messagebox.showinfo("Exportado", f"✅ Gráfico exportado com sucesso:\n{file_path}")
    except Exception as e:
        messagebox.showerror("Erro ao Exportar", f"❌ Ocorreu um erro ao salvar a imagem:\n{e}")



def gerar_grafico_reprovas_por_data():
    cursor.execute("SELECT data_reprova FROM produtos")
    dados = cursor.fetchall()
    if not dados:
        messagebox.showinfo("Sem dados", "Não há dados para gerar gráfico.")
        return
    datas = [row[0] for row in dados if row[0]]
    df = pd.DataFrame({'data': pd.to_datetime(datas, errors='coerce')})
    df.dropna(inplace=True)
    df['semana'] = df['data'].dt.to_period('W').astype(str)
    df['mes'] = df['data'].dt.to_period('M').astype(str)
    escolha = messagebox.askquestion("Escolha", "Deseja agrupar por mês? (Não = por semana)")
    agrupamento = 'mes' if escolha == 'yes' else 'semana'
    contagem = df[agrupamento].value_counts().sort_index()

    fig = px.bar(
        x=contagem.index, y=contagem.values,
        labels={'x': f"{agrupamento.capitalize()}", 'y': 'Qtd. de Reprovas'},
        text=contagem.values, title=f"Reprovas por {agrupamento.capitalize()}",
        color=contagem.index, color_discrete_sequence=px.colors.sequential.Reds
    )
    fig.update_traces(textposition='outside')
    fig.update_layout(template='plotly_white', showlegend=False, title_x=0.5)

    salvar_grafico_jpeg(fig, f"Reprovas_por_{agrupamento.capitalize()}")


def gerar_grafico_por_turno():
    cursor.execute("SELECT turno, COUNT(*) FROM produtos GROUP BY turno")
    dados = cursor.fetchall()
    if not dados:
        messagebox.showinfo("Sem dados", "Não há dados para gerar gráfico.")
        return
    turnos, quantidades = zip(*dados)

    fig = px.bar(
        x=turnos, y=quantidades,
        text=quantidades, title="Reprovas por Turno",
        labels={'x': 'Turno', 'y': 'Qtd. de Reprovas'},
        color=turnos, color_discrete_sequence=px.colors.sequential.Reds
    )
    fig.update_traces(textposition="outside")
    fig.update_layout(template="plotly_white", showlegend=False, title_x=0.5)

    salvar_grafico_jpeg(fig, "Reprovas_por_Turno")


def gerar_grafico_por_tipo_e_tag():
    cursor.execute("SELECT tipo_defeito, tag, COUNT(*) FROM produtos GROUP BY tipo_defeito, tag")
    dados = cursor.fetchall()
    if not dados:
        messagebox.showinfo("Sem dados", "Não há dados para gerar gráfico.")
        return
    df = pd.DataFrame(dados, columns=["Tipo Defeito", "Tag", "Qtd"])

    fig = px.bar(
        df, x="Tipo Defeito", y="Qtd", color="Tag", barmode="group",
        text="Qtd", title="Reprovas por Tipo de Defeito e Tag",
        color_discrete_sequence=px.colors.sequential.Reds
    )
    fig.update_traces(textposition="outside")
    fig.update_layout(template="plotly_white", title_x=0.5)

    salvar_grafico_jpeg(fig, "Tipo_Defeito_x_Tag")


def gerar_grafico_volume_por_tag():
    cursor.execute("SELECT tag, SUM(quantidade_total) FROM produtos GROUP BY tag")
    dados = cursor.fetchall()
    if not dados:
        messagebox.showinfo("Sem dados", "Não há dados para gerar o gráfico.")
        return

    tags = [linha[0] for linha in dados]
    volumes = [linha[1] for linha in dados]

    fig = px.pie(
        names=tags,
        values=volumes,
        title="Volume Total por Tag (Index / PIDM)",
        color_discrete_sequence=px.colors.sequential.Reds,
        hole=0.3
    )
    fig.update_traces(textposition="inside", textinfo="label+value")
    fig.update_layout(title_x=0.5, template="plotly_white")

    salvar_grafico_jpeg(fig, "Volume_Total_por_Tag")


def abrir_opcoes_graficos():
    global app
    janela = tk.Toplevel(app)
    janela.title("Escolha o Gráfico")
    janela.geometry("300x300")

    tk.Label(janela, text="Gráficos", font=("Arial", 12, "bold")).pack(pady=10)

    tk.Button(
        janela, text="1. Reprovas por Data", width=30,
        command=lambda: [janela.destroy(), gerar_grafico_reprovas_por_data()]
    ).pack(pady=5)

    tk.Button(
        janela, text="2. Reprovas por Turno", width=30,
        command=lambda: [janela.destroy(), gerar_grafico_por_turno()]
    ).pack(pady=5)

    tk.Button(
        janela, text="3. Tipo de Defeito x Tag", width=30,
        command=lambda: [janela.destroy(), gerar_grafico_por_tipo_e_tag()]
    ).pack(pady=5)

    tk.Button(
        janela, text="4. Volume Total por Tag", width=30,
        command=lambda: [janela.destroy(), gerar_grafico_volume_por_tag()]
    ).pack(pady=5)


from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def exportar_movimentacoes_excel():
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Arquivos Excel", "*.xlsx")],
        title="Salvar Movimentações como Excel"
    )
    if not file_path:
        return

    dados = [tabela_mov.item(row)["values"] for row in tabela_mov.get_children()]
    if not dados:
        messagebox.showinfo("Sem dados", "Não há movimentações para exportar.")
        return

    colunas = ["ID", "Produto ID", "Tipo", "Quantidade", "Data"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Movimentações"

    # Cabeçalho formatado (vermelho escuro e branco)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="EB1700")
    header_align = Alignment(horizontal="center")

    for col_index, titulo in enumerate(colunas, start=1):
        cell = ws.cell(row=1, column=col_index, value=titulo)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Inserir dados
    for row_index, linha in enumerate(dados, start=2):
        for col_index, valor in enumerate(linha, start=1):
            ws.cell(row=row_index, column=col_index, value=valor)

    # Largura automática das colunas
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        col_letra = get_column_letter(col[0].column)
        ws.column_dimensions[col_letra].width = max_len + 2

    wb.save(file_path)
    messagebox.showinfo("Exportado", "Movimentações exportadas com sucesso em Excel!")


def iniciar_app(user):
    global app, tabela, tabela_mov, tabela_log
    global codigo_entry, lote_entry, ordem_entry, qtd_total_entry, qtd_defeituosa_entry
    global tipo_defeito_entry, classe_defeito_entry, tag_entry, turno_entry
    global data_reprova_entry, tecnico_entry, comentarios_entry, foto_path
    global entradas, mov_produto_id, mov_tipo, mov_quantidade, cols
    global filtro_usuario, filtro_data

    app = tk.Tk()
    app.title("J&JControls")
    app.geometry("1100x800")

    notebook = ttk.Notebook(app)
    notebook.pack(fill="both", expand=True)

    # === Aba Cadastro ===
    aba_cadastro = tk.Frame(notebook)
    notebook.add(aba_cadastro, text="Cadastro de Produto")

    frame = tk.Frame(aba_cadastro)
    frame.pack(pady=10)

    campos = [
        ("Código do Produto", 0), ("Lote", 1), ("Ordem de Produção", 2), ("Qtd. Total", 3),
        ("Qtd. Defeituosa", 4), ("Tipo Defeito", 5), ("Classe Defeito", 6), ("Tag (INDEX/PIDM)", 7),
        ("Turno", 8), ("Data Reprova (DD/MM/AAAA)", 9), ("Técnico", 10), ("Comentários", 11)
    ]
    entradas = []

    for nome, i in campos:
        tk.Label(frame, text=nome).grid(row=i, column=0, padx=5, pady=2, sticky="w")
        entry = tk.Entry(frame, width=40)
        entry.grid(row=i, column=1, padx=5, pady=2)
        entradas.append(entry)

    (codigo_entry, lote_entry, ordem_entry, qtd_total_entry, qtd_defeituosa_entry,
     tipo_defeito_entry, classe_defeito_entry, tag_entry, turno_entry,
     data_reprova_entry, tecnico_entry, comentarios_entry) = entradas

    foto_path = tk.StringVar()
    tk.Button(frame, text="Anexar Foto da Reprova", command=selecionar_foto).grid(row=13, column=0, columnspan=2, pady=5)
    tk.Button(frame, text="Salvar Produto", command=salvar_produto).grid(row=14, column=0, columnspan=2, pady=10)

    busca_entry = tk.Entry(aba_cadastro)
    busca_entry.pack(pady=5)
    tk.Button(aba_cadastro, text="Filtrar", command=lambda: atualizar_tabela(busca_entry.get())).pack(pady=5)
    tk.Button(aba_cadastro, text="Exportar Excel", command=exportar_produtos_excel).pack(pady=2)

    cols = ("ID", "Código", "Lote", "Ordem", "Qtd Total", "Qtd Defeituosa", "Tipo Defeito",
            "Classe Defeito", "Tag", "Turno", "Data Reprova", "Técnico", "Comentários", "Foto")
    tabela = ttk.Treeview(aba_cadastro, columns=cols, show="headings", height=10)
    for col in cols:
        tabela.heading(col, text=col)
        tabela.column(col, width=90)
    tabela.pack(fill="x")

    # === Aba Movimentações ===
    aba_mov = tk.Frame(notebook)
    notebook.add(aba_mov, text="Movimentações")

    tk.Label(aba_mov, text="ID do Produto").pack(pady=2)
    mov_produto_id = tk.Entry(aba_mov)
    mov_produto_id.pack(pady=2)

    tk.Label(aba_mov, text="Tipo (Entrada/Saída)").pack(pady=2)
    mov_tipo = ttk.Combobox(aba_mov, values=["Entrada", "Saída"])
    mov_tipo.pack(pady=2)

    tk.Label(aba_mov, text="Quantidade").pack(pady=2)
    mov_quantidade = tk.Entry(aba_mov)
    mov_quantidade.pack(pady=2)

    tk.Button(aba_mov, text="Registrar Movimentação", command=registrar_movimentacao).pack(pady=10)

    cols_mov = ("ID", "Produto ID", "Tipo", "Quantidade", "Data")
    tabela_mov = ttk.Treeview(aba_mov, columns=cols_mov, show="headings", height=10)
    for col in cols_mov:
        tabela_mov.heading(col, text=col)
        tabela_mov.column(col, width=100)
    tabela_mov.pack(fill="x")
    tk.Button(aba_mov, text="Exportar Excel", command=exportar_movimentacoes_excel).pack(pady=5)

        # === Aba Histórico de Operações ===
    aba_log = tk.Frame(notebook)
    notebook.add(aba_log, text="Histórico de Operações")

    frame_filtros = tk.Frame(aba_log)
    frame_filtros.pack(pady=5)

    tk.Label(frame_filtros, text="Usuário:").grid(row=0, column=0, padx=5)
    filtro_usuario = tk.Entry(frame_filtros)
    filtro_usuario.grid(row=0, column=1, padx=5)

    tk.Label(frame_filtros, text="Data (DD/MM/AAAA):").grid(row=0, column=2, padx=5)
    filtro_data = tk.Entry(frame_filtros)
    filtro_data.grid(row=0, column=3, padx=5)

    def atualizar_tabela_log():
        for row in tabela_log.get_children():
            tabela_log.delete(row)
        cursor.execute("SELECT * FROM log_operacoes ORDER BY id DESC")
        for row in cursor.fetchall():
            tabela_log.insert("", tk.END, values=row)

    def filtrar_log():
        usuario = filtro_usuario.get().strip()
        data = filtro_data.get().strip()

        query = "SELECT * FROM log_operacoes WHERE 1=1"
        params = []

        if usuario:
            query += " AND usuario LIKE ?"
            params.append(f"%{usuario}%")

        if data:
            try:
                data_formatada = datetime.strptime(data, "%d/%m/%Y").strftime("%Y-%m-%d")
                query += " AND DATE(data_hora) = ?"
                params.append(data_formatada)
            except ValueError:
                messagebox.showwarning("Data inválida", "Use o formato DD/MM/AAAA.")
                return

        for row in tabela_log.get_children():
            tabela_log.delete(row)

        cursor.execute(query, params)
        for row in cursor.fetchall():
            tabela_log.insert("", tk.END, values=row)

    def exportar_log_excel():
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Arquivos Excel", "*.xlsx")],
            title="Salvar Log como Excel"
        )
        if not file_path:
            return

        dados = [tabela_log.item(row)["values"] for row in tabela_log.get_children()]
        if not dados:
            messagebox.showinfo("Sem dados", "Não há registros para exportar.")
            return

        colunas = ["ID", "Operação", "Código", "Lote", "Usuário", "Data/Hora"]
        wb = Workbook()
        ws = wb.active
        ws.title = "Histórico"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="EB1700")
        header_align = Alignment(horizontal="center")

        for col_index, titulo in enumerate(colunas, start=1):
            cell = ws.cell(row=1, column=col_index, value=titulo)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for row_index, linha in enumerate(dados, start=2):
            for col_index, valor in enumerate(linha, start=1):
                ws.cell(row=row_index, column=col_index, value=valor)

        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            col_letra = get_column_letter(col[0].column)
            ws.column_dimensions[col_letra].width = max_len + 2

        wb.save(file_path)
        messagebox.showinfo("Exportado", "Log exportado com sucesso em Excel!")

    # Botões de ações
    tk.Button(frame_filtros, text="Filtrar", command=filtrar_log).grid(row=0, column=4, padx=10)
    tk.Button(frame_filtros, text="Atualizar Log", command=atualizar_tabela_log).grid(row=0, column=5, padx=5)
    tk.Button(frame_filtros, text="Exportar Log Excel", command=exportar_log_excel).grid(row=0, column=6, padx=5)

    # Tabela de log
    cols_log = ("ID", "Operação", "Código", "Lote", "Usuário", "Data/Hora")
    tabela_log = ttk.Treeview(aba_log, columns=cols_log, show="headings", height=10)
    for col in cols_log:
        tabela_log.heading(col, text=col)
        tabela_log.column(col, width=120)
    tabela_log.pack(fill="x")

    atualizar_tabela_log()

            # === Aba Gráficos ===
    aba_graficos = tk.Frame(notebook)
    notebook.add(aba_graficos, text="Gráficos")

    tk.Label(aba_graficos, text="Visualização de Gráficos", font=("Arial", 12, "bold")).pack(pady=10)

    tk.Button(aba_graficos, text="1. Reprovas por Data", width=30,
              command=gerar_grafico_reprovas_por_data).pack(pady=5)

    tk.Button(aba_graficos, text="2. Reprovas por Turno", width=30,
              command=gerar_grafico_por_turno).pack(pady=5)

    tk.Button(aba_graficos, text="3. Tipo de Defeito x Tag", width=30,
              command=gerar_grafico_por_tipo_e_tag).pack(pady=5)
    
    tk.Button(aba_graficos, text="4. Volume Total por Tag", width=30,
              command=gerar_grafico_volume_por_tag).pack(pady=5)


    
def exportar_log_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Arquivos Excel", "*.xlsx")],
        title="Salvar Log como Excel"
    )
    if not file_path:
        return

    # Capturar dados da tabela
    dados = [tabela_log.item(row)["values"] for row in tabela_log.get_children()]
    if not dados:
        messagebox.showinfo("Sem dados", "Não há registros para exportar.")
        return

    colunas = ["ID", "Operação", "Código", "Lote", "Usuário", "Data/Hora"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Histórico"

    # Estilo do cabeçalho
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="EB1700")
    header_align = Alignment(horizontal="center")

    for col_index, titulo in enumerate(colunas, start=1):
        cell = ws.cell(row=1, column=col_index, value=titulo)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Inserir os dados
    for row_index, linha in enumerate(dados, start=2):
        for col_index, valor in enumerate(linha, start=1):
            ws.cell(row=row_index, column=col_index, value=valor)

    # Ajustar largura automática das colunas
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        col_letra = get_column_letter(col[0].column)
        ws.column_dimensions[col_letra].width = max_len + 2

    wb.save(file_path)
    messagebox.showinfo("Exportado", "Log exportado com sucesso em Excel!")


# ====== INICIAR COM LOGIN ======
if __name__ == "__main__":
    abrir_login(iniciar_app)
